import tkinter as tk
import re
import openpyxl
from tkinter import font
from datetime import datetime

# Función para validar la longitud de una cadena
def validar_longitud(cadena, limite):
    return len(cadena) <= limite

# Función para validar el formato de un correo electrónico utilizando expresiones regulares
def validar_correo(email):
    email_regex = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return re.match(email_regex, email) is not None

# Función para validar los datos ingresados en el formulario
def validar_datos():
    usuario = entry_usuario.get()
    pass1 = entry_pass1.get()
    pass2 = entry_pass2.get()
    email = entry_email.get()

    # Validar la longitud del nombre de usuario
    if not validar_longitud(usuario, 10):
        resultado_label.config(text="La longitud del nombre de usuario debe ser 10 caracteres", fg="red")
    # Validar la longitud de las contraseñas y su correspondencia
    elif not validar_longitud(pass1, 8) or not validar_longitud(pass2, 8):
        resultado_label.config(text="La longitud de la contraseña debe ser 8 caracteres", fg="red")
    elif not re.search('[0-9]', pass1) or pass1 != pass2:
        resultado_label.config(text="La contraseña es incorrecta", fg="red")
    # Validar el formato del correo electrónico
    elif not validar_correo(email):
        resultado_label.config(text="El email es incorrecto", fg="red")
    else:
        guardar_datos(usuario, pass1, email)
        resultado_label.config(text="Datos Correctos", fg="green")

# Función para guardar los datos en el archivo datos_usuarios.xlsx
def guardar_datos(usuario, pass1, email):
    try:
        workbook = openpyxl.load_workbook('datos_usuarios.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.save('datos_usuarios.xlsx')

    # Comprobar si existe la hoja 'Usuarios'; si no, crearla con encabezados
    if 'Usuarios' in workbook.sheetnames:
        sheet = workbook['Usuarios']
    else:
        sheet = workbook.create_sheet('Usuarios')
        sheet.append(["ID", "Usuario", "Contraseña", "Email", "Fecha de Registro"])  # Añadir encabezados si la hoja es nueva

    ultima_fila = sheet.max_row
    contador = str(ultima_fila).zfill(3)  # Formatear el contador como una cadena de tres dígitos

    fecha_registro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Añadir los datos del usuario a la hoja
    sheet.append([contador, usuario, pass1, email, fecha_registro])

    workbook.save('datos_usuarios.xlsx')

    # Limpiar los campos del formulario después de guardar los datos
    entry_usuario.delete(0, tk.END)
    entry_pass1.delete(0, tk.END)
    entry_pass2.delete(0, tk.END)
    entry_email.delete(0, tk.END)

# Crear la ventana principal y configurarla
window = tk.Tk()
window.title("Formulario de Registro")
window.geometry("420x250")

bg_color = "#f0f0f0"
entry_bg_color = "white"
button_bg_color = "#4CAF50"
button_fg_color = "white"
bold_font = font.Font(weight="bold")

# Crear etiquetas y campos de entrada para el formulario
label_usuario = tk.Label(window, text="Nombre de usuario:", bg=bg_color, font=bold_font)
label_usuario.grid(row=0, column=0, sticky="w", padx=10, pady=5)
entry_usuario = tk.Entry(window, bg=entry_bg_color, width=25)
entry_usuario.grid(row=0, column=1, padx=10, pady=5)

label_pass1 = tk.Label(window, text="Contraseña:", bg=bg_color, font=bold_font)
label_pass1.grid(row=1, column=0, sticky="w", padx=10, pady=5)
entry_pass1 = tk.Entry(window, bg=entry_bg_color, width=25, show="*")
entry_pass1.grid(row=1, column=1, padx=10, pady=5)

label_pass2 = tk.Label(window, text="Confirmar Contraseña:", bg=bg_color, font=bold_font)
label_pass2.grid(row=2, column=0, sticky="w", padx=10, pady=5)
entry_pass2 = tk.Entry(window, bg=entry_bg_color, width=25, show="*")
entry_pass2.grid(row=2, column=1, padx=10, pady=5)

label_email = tk.Label(window, text="Correo electrónico:", bg=bg_color, font=bold_font)
label_email.grid(row=3, column=0, sticky="w", padx=10, pady=5)
entry_email = tk.Entry(window, bg=entry_bg_color, width=25)
entry_email.grid(row=3, column=1, padx=10, pady=5)

# Crear botón de enviar
submit_button = tk.Button(window, text="Enviar", bg=button_bg_color, fg=button_fg_color, font=bold_font, command=validar_datos)
submit_button.grid(row=4, column=0, columnspan=2, pady=10)

# Crear etiqueta para mostrar el resultado de la validación
resultado_label = tk.Label(window, text="", bg=bg_color, font=bold_font)
resultado_label.grid(row=5, column=0, columnspan=2)

# Iniciar el bucle de eventos
window.mainloop()

